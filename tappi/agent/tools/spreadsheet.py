"""Spreadsheet tool — CSV and Excel read/write.

CSV: stdlib csv module (zero deps).
Excel: openpyxl (optional dep).
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from tappi.agent.config import get_workspace

TOOL_SCHEMA = {
    "type": "function",
    "function": {
        "name": "spreadsheet",
        "description": (
            "Read and write CSV and Excel (.xlsx) files. Can read data, "
            "write rows, create new spreadsheets, and query specific columns or ranges."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["read", "write", "create", "info"],
                    "description": (
                        "Spreadsheet action:\n"
                        "- read: Read contents (requires 'path', optional 'sheet', 'rows', 'columns')\n"
                        "- write: Append rows to existing file (requires 'path' and 'rows')\n"
                        "- create: Create new file with headers and data (requires 'path', 'headers', optional 'rows')\n"
                        "- info: Get sheet names, row/column counts (requires 'path')"
                    ),
                },
                "path": {"type": "string", "description": "File path (relative to workspace). .csv or .xlsx"},
                "sheet": {"type": "string", "description": "Sheet name for Excel files (default: active sheet)"},
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column headers for create action",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array", "items": {"type": "string"}},
                    "description": "Data rows (array of arrays) for write/create",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column names to include in read (default: all)",
                },
                "max_rows": {"type": "integer", "description": "Max rows to read (default: 500)"},
            },
            "required": ["action", "path"],
        },
    },
}


class SpreadsheetTool:
    """CSV and Excel operations, sandboxed to workspace."""

    def __init__(self, workspace: Path | None = None) -> None:
        self._workspace = workspace

    @property
    def workspace(self) -> Path:
        if self._workspace is None:
            self._workspace = get_workspace()
        self._workspace = self._workspace.resolve()
        self._workspace.mkdir(parents=True, exist_ok=True)
        return self._workspace

    def _resolve(self, path: str) -> Path:
        resolved = (self.workspace / path).resolve()
        ws_resolved = self.workspace.resolve()
        if not str(resolved).startswith(str(ws_resolved)):
            raise PermissionError(f"Access denied: path escapes workspace")
        return resolved

    def _is_excel(self, path: str) -> bool:
        return path.lower().endswith((".xlsx", ".xls"))

    def execute(self, **params: Any) -> str:
        action = params.get("action", "")
        try:
            if action == "read":
                return self._read(params)
            elif action == "write":
                return self._write(params)
            elif action == "create":
                return self._create(params)
            elif action == "info":
                return self._info(params)
            else:
                return f"Unknown action: {action}"
        except ImportError:
            return "Missing dependency: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"

    def _read(self, params: dict) -> str:
        path = params.get("path", "")
        resolved = self._resolve(path)
        if not resolved.exists():
            return f"File not found: {path}"

        max_rows = int(params.get("max_rows", 500))
        filter_cols = params.get("columns")

        if self._is_excel(path):
            return self._read_excel(resolved, params.get("sheet"), filter_cols, max_rows)
        else:
            return self._read_csv(resolved, filter_cols, max_rows)

    def _read_csv(self, resolved: Path, filter_cols: list | None, max_rows: int) -> str:
        with open(resolved, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            if filter_cols:
                headers = [h for h in headers if h in filter_cols]

            lines = [",".join(headers)]
            count = 0
            for row in reader:
                if count >= max_rows:
                    lines.append(f"\n... (truncated at {max_rows} rows)")
                    break
                vals = [str(row.get(h, "")) for h in headers]
                lines.append(",".join(vals))
                count += 1

        return "\n".join(lines)

    def _read_excel(self, resolved: Path, sheet: str | None, filter_cols: list | None, max_rows: int) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(str(resolved), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return "(empty spreadsheet)"

        headers = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
        if filter_cols:
            col_indices = [i for i, h in enumerate(headers) if h in filter_cols]
            headers = [headers[i] for i in col_indices]
        else:
            col_indices = list(range(len(headers)))

        lines = [",".join(headers)]
        for row in rows[1:max_rows + 1]:
            vals = [str(row[i]) if i < len(row) and row[i] is not None else "" for i in col_indices]
            lines.append(",".join(vals))

        if len(rows) - 1 > max_rows:
            lines.append(f"\n... (truncated at {max_rows} rows, {len(rows) - 1} total)")

        return "\n".join(lines)

    def _write(self, params: dict) -> str:
        path = params.get("path", "")
        rows = params.get("rows", [])
        if not rows:
            return "Error: 'rows' required for write action."

        resolved = self._resolve(path)
        if not resolved.exists():
            return f"File not found: {path}. Use action='create' to make a new file."

        if self._is_excel(path):
            from openpyxl import load_workbook
            wb = load_workbook(str(resolved))
            ws = wb.active
            for row in rows:
                ws.append([str(v) for v in row])
            wb.save(str(resolved))
            wb.close()
        else:
            with open(resolved, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in rows:
                    writer.writerow(row)

        return f"Appended {len(rows)} rows to {path}"

    def _create(self, params: dict) -> str:
        path = params.get("path", "")
        headers = params.get("headers", [])
        rows = params.get("rows", [])

        resolved = self._resolve(path)
        resolved.parent.mkdir(parents=True, exist_ok=True)

        if self._is_excel(path):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            if headers:
                ws.append(headers)
            for row in rows:
                ws.append([str(v) for v in row])
            wb.save(str(resolved))
            wb.close()
        else:
            with open(resolved, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                for row in rows:
                    writer.writerow(row)

        total = len(rows) + (1 if headers else 0)
        return f"Created: {path} ({total} rows)"

    def _info(self, params: dict) -> str:
        path = params.get("path", "")
        resolved = self._resolve(path)
        if not resolved.exists():
            return f"File not found: {path}"

        if self._is_excel(path):
            from openpyxl import load_workbook
            wb = load_workbook(str(resolved), read_only=True)
            sheets = wb.sheetnames
            ws = wb.active
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            wb.close()
            return f"File: {path}\nType: Excel\nSheets: {', '.join(sheets)}\nRows: {rows}\nColumns: {cols}"
        else:
            with open(resolved, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows_list = list(reader)
            headers = rows_list[0] if rows_list else []
            return f"File: {path}\nType: CSV\nHeaders: {', '.join(headers)}\nRows: {len(rows_list) - 1}"
